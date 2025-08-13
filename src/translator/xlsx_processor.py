import asyncio
from openpyxl import load_workbook, Workbook
from typing import List, Tuple, Any
from .llm_client import DeepSeekClient

class XlsxProcessor:
    def __init__(self):
        self.workbook = None
        
    def load_file(self, file_path: str):
        self.workbook = load_workbook(file_path)
        
    def get_all_text_cells(self) -> List[Tuple[Any, str]]:
        cells_data = []
        for sheet in self.workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.strip():
                        cells_data.append((cell, cell.value))
        return cells_data
    
    async def translate_xlsx(self, file_path: str, target_lang: str, max_concurrent: int = 50,
                           progress_callback=None, log_callback=None) -> str:
        if log_callback:
            log_callback("Loading Excel file...")
        
        self.load_file(file_path)
        cells_data = self.get_all_text_cells()
        
        if not cells_data:
            if log_callback:
                log_callback("No text content found in file")
            return file_path
            
        total_cells = len(cells_data)
        if log_callback:
            log_callback(f"Found {total_cells} text cells, starting translation...")
            
        texts = [cell_data[1] for cell_data in cells_data]
        
        async with DeepSeekClient() as client:
            if log_callback:
                log_callback("Connecting to DeepSeek API...")
            translated_texts = await client.batch_translate(
                texts, target_lang, max_concurrent,
                progress_callback=lambda i: (
                    progress_callback and progress_callback(i, total_cells),
                    log_callback and log_callback(f"Translated {i}/{total_cells} cells")
                )
            )
            
        if log_callback:
            log_callback("Updating Excel file...")
            
        # 更新单元格
        for (cell, _), translated_text in zip(cells_data, translated_texts):
            cell.value = translated_text
        
        # 保存文件
        output_path = file_path.replace('.xlsx', f'_translated_{target_lang}.xlsx')
        self.workbook.save(output_path)
        
        if log_callback:
            log_callback(f"Translation completed! File saved: {output_path}")
            
        return output_path